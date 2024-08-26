from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
import openpyxl as excel
from search.models import Projectinfo
import re
from django.views.generic import DetailView

#案件・技術者検索追加
from typing import Any
from django.db.models.query import QuerySet
from django.db.models import Q
import re
from django.views.generic import ListView
from search.models import Projectinfo,Engineerinfo,Skillinfo,Codemaster

from django.views.generic import ListView
from .models import Engineerinfo



######################################
#トップ画面
#######################################

def top(request):
     return render(request, 'search/top.html')


######################################
#案件情報登録処理
#######################################

def basic_upload(request):
    # 画面上で受け取ったファイルを受け取る
    if request.method == 'POST' and request.FILES['testfile']:
        myfile = request.FILES['testfile']
        
        #　Excel以外のファイルが渡されたときにエラー文を出力するための処理
        if not myfile.name.endswith(".xlsx"):
            error = "Excelファイル以外のファイルが選択されました。"
            context = {"error" : error
            }
            return render(request, 'search/project_file_upload.html', context)

        fs = FileSystemStorage()
        fs.save(myfile.name, myfile)

        try:
            #ワークブックを開く
            book = excel.load_workbook(myfile, data_only=True)
            #案件情報一覧.xlsxの最初(0番目)のシートを取り出す
            sheet = book.worksheets[0]

            # Excelに記載されてる回数処理を行う
            for row in sheet.iter_rows(min_row=2):
                if row[0].value is None:
                    return render(request, 'search/error.html')
                # 案件情報をリストに格納する
                project_list = []
                for project_data in row:
                    project_list.append(project_data.value)
                
                # 各変数に取得した案件情報を格納する
                # ※projectinfoのみ多くの情報が含まれる文字列で格納されている
                jobtype = project_list[1]
                recruitmentnumbers = project_list[2]
                period = project_list[3]
                clientname = project_list[4]
                projectinfo = project_list[5]            
                remarks = project_list[6]

                # 変数projectinfoがNoneでなかった場合処理を行う
                if not projectinfo == None:
                    # '【'で文字列を分割
                    split = re.split('【', projectinfo)

                    # 変数splitの配列の長さ分処理を行う
                    # ---以下for文の中の処理内容---
                    # 1. split[i]の内容を各変数に格納
                    # 2. '〇〇】'で文字列を再度分割
                    # 3. 改行の回数をcount変数に格納
                    # 4. 改行の回数-1回分改行を空白に変換し、全ての空白をなしにする。それを各カラムの変数に格納
                    workcontents = ""
                    workprocess = ""
                    mustskill = ""
                    passableskill = ""
                    skilllevel = ""
                    paymentunitprice = ""
                    numberofinterviews = ""
                    foreignnationality = ""
                    bpallowed = ""
                    startdate = ""
                    neareststation = ""
                    telework = ""
                    for i in range(len(split)):
                        if '作業内容' in split[i] :
                            workcontents = split[i]
                            workcontents = re.split('作業内容】', workcontents)
                            count = workcontents[1].count("\n")
                            workcontents = workcontents[1].replace("\n", " ",count-1).replace("\u3000", "")
                        elif '作業工程' in split[i] :
                            workprocess = split[i]
                            workprocess = re.split('作業工程】', workprocess)
                            count = workprocess[1].count("\n")
                            workprocess = workprocess[1].replace("\n", " ",count-1).replace("\u3000", "")
                        #　案件魅力は案件一覧の項目にあるもののDBにはないためコメントアウト 
                        # elif '案件魅力' in split[i] :
                        #   a = split[i]
                        #   a = re.split('案件魅力】', a) 
                        #   count = a[1].count("\n")
                        #   a = a[1].replace("\n", " ",count-1).replace("\u3000", "")         
                        elif 'Must' in split[i] :
                            mustskill = split[i]
                            mustskill = re.split('Must】', mustskill)
                            count = mustskill[1].count("\n")
                            mustskill = mustskill[1].replace("\n", " ",count-1).replace("\u3000", "")
                        elif 'Want' in split[i] :
                            passableskill = split[i]
                            passableskill = re.split('Want】', passableskill)
                            count = passableskill[1].count("\n")
                            passableskill = passableskill[1].replace("\n", " ",count-1).replace("\u3000", "") 
                        elif 'レベル' in split[i] :
                            skilllevel = split[i]
                            skilllevel = re.split('レベル】', skilllevel)
                            count = skilllevel[1].count("\n")
                            skilllevel = skilllevel[1].replace("\n", " ",count-1).replace("\u3000", "")
                        elif '支払単価' in split[i] :
                            paymentunitprice = split[i]
                            paymentunitprice = re.split('支払単価】', paymentunitprice)
                            count = paymentunitprice[1].count("\n")
                            paymentunitprice = paymentunitprice[1].replace("\n", " ",count-1).replace("\u3000", "") 
                        elif '面談回数' in split[i] :
                            numberofinterviews = split[i]
                            numberofinterviews = re.split('面談回数】', numberofinterviews)
                            count = numberofinterviews[1].count("\n")
                            numberofinterviews = numberofinterviews[1].replace("\n", " ",count-1).replace("\u3000", "") 
                        elif '外国籍' in split[i] :
                            foreignnationality = split[i]
                            foreignnationality = re.split('外国籍】', foreignnationality)
                            count = foreignnationality[1].count("\n")
                            foreignnationality = foreignnationality[1].replace("\n", " ",count-1).replace("\u3000", "") 
                        elif '1社先BP' in split[i] :
                            bpallowed = split[i]
                            bpallowed = re.split('1社先BP】', bpallowed)
                            count = bpallowed[1].count("\n")
                            bpallowed = bpallowed[1].replace("\n", " ",count-1).replace("\u3000", "")  
                        elif '開始時期' in split[i] :
                            startdate = split[i]
                            startdate = re.split('開始時期】', startdate)
                            count = startdate[1].count("\n")
                            startdate = startdate[1].replace("\n", " ",count-1).replace("\u3000", "")  
                        # elif '募集人数' in split[i] :
                        #     recruitmentnumbers = split[i]
                        #     recruitmentnumbers = re.split('募集人数】', recruitmentnumbers)
                        #     count = recruitmentnumbers[1].count("\n")
                        #     recruitmentnumbers = recruitmentnumbers[1].replace("\n", " ",count-1).replace("\u3000", "")  
                        elif '最寄駅' in split[i] :
                            neareststation = split[i]
                            neareststation = re.split('最寄駅】', neareststation)
                            count = neareststation[1].count("\n")
                            neareststation = neareststation[1].replace("\n", " ",count-1).replace("\u3000", "")  
                        elif 'テレワーク' in split[i] :
                            telework = split[i]
                            telework = re.split('テレワーク】', telework)
                            count = telework[1].count("\n")
                            telework = telework[1].replace("\n", " ",count-1).replace("\u3000", "")  

                    # 各変数をDBに登録
                    insert = Projectinfo(jobtype=jobtype, recruitmentnumbers=recruitmentnumbers, period=period, clientname=clientname, workcontents=workcontents, workprocess=workprocess, mustskill=mustskill, passableskill=passableskill, skilllevel=skilllevel, paymentunitprice=paymentunitprice, numberofinterviews=numberofinterviews, foreignnationality=foreignnationality, bpallowed=bpallowed, startdate=startdate, neareststation=neareststation, telework=telework, remarks=remarks)
                    insert.save()

                # 変数projectinfoがNoneの時に各変数をDBに登録
                else:
                    insert = Projectinfo(jobtype=jobtype, recruitmentnumbers=recruitmentnumbers, period=period, clientname=clientname, remarks=remarks)
                    insert.save()

            # ロードしたExcelファイルを閉じる
            book.close()
            complete = {"complete" : True
            }
            return render(request, 'search/project_file_upload.html', complete)
        
        except:
            return render(request, 'search/error.html')
        

    return render(request, 'search/project_file_upload.html')


#######################################
#技術者検索、案件情報検索処理
#######################################

class EngineerListView(ListView):
    model = Engineerinfo
    template_name = "search/search_engineer_list.html"
    context_object_name = "engineers"
    paginate_by = 12

class SearchSkillEngineerListView(ListView):
    template_name = "search/search_engineer_list.html"
    context_object_name = 'engineers'
    paginate_by = 12

    def get_queryset(self):
        skill_query = self.request.GET.getlist('skills')  # チェックされたスキルのリストを取得

        if skill_query:
            engineer_ids = None
            for skill in skill_query:
                skill_engineer_ids = Skillinfo.objects.filter(classification__icontains=skill).values_list('id', flat=True).distinct()

                if engineer_ids is None:
                    engineer_ids = set(skill_engineer_ids)
                else:
                    engineer_ids &= set(skill_engineer_ids)

            # engineer_idsがNoneまたは空の場合、空のクエリセットを返す
            if not engineer_ids:
                return Skillinfo.objects.none()
            # Engineerinfoテーブルから技術者を絞り込み
            queryset = Engineerinfo.objects.filter(id__in=engineer_ids)
        else:
            queryset = Engineerinfo.objects.none()

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # skills というキーでスキルのリストを取得する
        context['skills'] = Codemaster.objects.all()
        
        return context


class SearchNameEngineerListView(ListView):
    model = Engineerinfo
    template_name = "search/search_engineer_list.html"
    context_object_name = "engineers"
    paginate_by = 12

    def get_queryset(self):
        self.query = self.request.GET.get("query") or ""
        queryset= super().get_queryset()

        if self.query:
            queryset = queryset.filter(Q(name__icontains=self.query))
            return queryset
       

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.query
        return context

class ProjectListView(ListView):
    model = Projectinfo
    template_name = 'search/search_project_list.html'
    context_object_name = "projects"
    paginate_by = 12
    
class SearchProjectByEngineerName(ListView):
    template_name = 'search/search_project_list.html'
    context_object_name = "projects"
    paginate_by = 12

    def get_queryset(self):
        query = self.request.GET.get("query")
        
        skill_classifications = None
        project_ids = None
        if query:
            # フォームが入力されている場合、技術者名によるフィルタリングを行う
            skillnames = Skillinfo.objects.filter(code__icontains=query).values_list("classification",flat=True).distinct()
            
            skill_classifications = set(skillnames)
            if not skill_classifications:
                return Projectinfo.objects.none()
            for skill_classification in skill_classifications:

                projectids=Projectinfo.objects.filter(Q(mustskill__icontains=skill_classification)).values_list("id",flat=True)
                if project_ids is None:
                    project_ids = set(projectids)
                else:
                    project_ids |= set(projectids)
            if not project_ids:
                return Projectinfo.objects.none()
            queryset = Projectinfo.objects.filter(id__in=project_ids)
        else:
            queryset = Projectinfo.objects.all()

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get("query")
        return context


class SearchSomethingProjectListView(ListView):
    
    template_name = "search/search_project_list.html"
    context_object_name = "projects"
    paginate_by = 12

    def get_queryset(self):
       
        queryset = Projectinfo.objects.all()
        projectname = self.request.GET.get('project_name') or ""
        freeword = self.request.GET.get('freeword') or ""
        period = self.request.GET.get('period') or ""
        skill_query = self.request.GET.getlist('skills')
        ids = None
        if (not projectname) & (not freeword) & (not period) & (not skill_query):
            return Projectinfo.objects.none()
        
        if projectname:
            queryset = queryset.filter(clientname__icontains=projectname)
        elif not projectname:
            queryset = queryset
        else:
            queryset = Projectinfo.objects.none()
            return queryset

        if freeword:
            freewords = re.split('[ ,　、]+',freeword)
            for k in freewords:
                projectids = queryset.filter(
                    Q(jobtype__icontains = k)|
                    Q(recruitmentnumbers__icontains = k)|
                    Q(workcontents__icontains = k) |
                    Q(workprocess__icontains = k)|
                    Q(mustskill__icontains = k)|
                    Q(passableskill = k)|
                    Q(skilllevel__icontains=k)|
                    Q(paymentunitprice__icontains = k)|
                    Q(numberofinterviews__icontains = k)|
                    Q(foreignnationality__icontains = k)|
                    Q(bpallowed__icontains = k)|
                    Q(startdate__icontains = k)|
                    Q(neareststation__icontains = k)|
                    Q(telework__icontains = k)|
                    Q(remarks__icontains = k)).values_list('id',flat=True)
                
                if ids is None:
                    ids = set(projectids)
                else:
                    ids &= set(projectids)

            queryset = queryset.filter(id__in = ids)
        
        elif not freeword:
            queryset = queryset
        
        else:
            queryset = Projectinfo.objects.none()
            return queryset
        

        if period:
            queryset = queryset.filter(period__icontains = period)

        elif not period:
            queryset = queryset.filter(period__icontains = period)
        
        else:
            return Projectinfo.objects.none()
        
        if skill_query:
            project_ids =None
            for skill in skill_query:
                skill_project_ids = queryset.filter(mustskill__icontains = skill).values_list("id",flat=True).distinct()

                if project_ids is None:
                    project_ids = set(skill_project_ids)
                else:
                    project_ids &= set(skill_project_ids)

            if not project_ids:
                return Projectinfo.objects.none()
            queryset = queryset.filter(id__in = project_ids)
        
        else:
            return queryset
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['project_name'] = self.request.GET.get('project_name') or ""
        context['freeword'] = self.request.GET.get('freeword') or ""
        context['period'] = self.request.GET.get('period', '')
        context['skills'] = Codemaster.objects.all()
        return context
    
#######################################
#技術者情報詳細処理
#######################################
class EngineerDetailView(DetailView):
    model = Engineerinfo
    template_name = 'search/engineer_detail_list.html'

    def get_context_data(self, **kwargs):
        project = []
        project_id = None
        code = None
        number = self.kwargs['pk']

        # 技術者情報の取得
        # idは受け取った値が入るので変更予定
        engineer = Engineerinfo.objects.get(id=number)
        
        #技術者の名前と一致するスキル情報をスキル情報テーブルから取得
        engineerskill = Skillinfo.objects.filter(code__iexact=engineer.name).values_list('classification',flat=True)
        
        # 案件情報の必須スキルに大文字・小文字関係なく文字列を含んでいた場合変数checkに格納
        # project_idがNoneまたはproject_idに↑で格納されたcheckの値が入っていない場合project_idにcheckを格納
        for engineerskill in engineerskill:
            check = Projectinfo.objects.filter(mustskill__icontains=engineerskill).values_list('id', flat=True).distinct()
            
            if project_id is None:
                project_id = set(check)
            else:
                project_id |= set(check)  
                
        # project_idと案件情報テーブルのidが同じ場合projectに格納
        if project_id != None:
            for project_id in project_id:
                check = Projectinfo.objects.filter(id=project_id)
                project.append(check)

        engineerskill_classification = Skillinfo.objects.filter(code__iexact=engineer.name).values_list('classification',flat=True)

        #技術者のスキル名と一致するスキルのコードマスタ情報をcodemaster_listに格納
        for engineerskill_classification in engineerskill_classification:
            check = Codemaster.objects.filter(name__iexact=engineerskill_classification)
            
            if code is None:
                code = set(check)
            else:
                code |= set(check)

        # contextに文字を入れて送る
        context = {'engineer' : engineer,
                    'project' : project,
                    'skilllist' : engineerskill,
                    'codemaster' : code
                }
        return context


#######################################
#案件情報詳細処理
#######################################

class ProjectDetailView(DetailView):
    model = Projectinfo
    template_name = 'search/project_detail_list.html'

    def get_context_data(self, **kwargs):
        engineer_list = []
        codemaster_lower = []
        skill = [] 
        engineer_list_ids = None
        code = None
        project_id = self.kwargs['pk']

        # 案件リストの取得
        # idは受け取った値が入るので変更予定
        project = Projectinfo.objects.get(id=project_id)

        # 案件情報の必須スキルを小文字に変換し、projectmustskill_lowerに格納
        if project.mustskill != None:
            projectmustskill_lower = project.mustskill.lower()

        # コードマスタのスキル名を小文字に変換し、codemaster_lowerに格納
            codemaster = Codemaster.objects.all()
            for codemaster in codemaster:
                change = codemaster.name.lower()
                codemaster_lower.append(change)
            
            # 案件情報の必須スキルに記載されている文章からコードマスタに登録されているスキル名が一致しているものを抜き出してskillに格納。
            for codemaster_lower in codemaster_lower:
                if codemaster_lower in projectmustskill_lower: 
                    change = codemaster_lower
                    skill.append(change)        

            # 案件情報の必須スキルと一致するスキルを有するスキルテーブルの技術者の技術者情報をcodeに格納  
            for skill in skill:
                check = Skillinfo.objects.filter(classification__icontains=skill)
                
                if code is None:
                    code = set(check)
                else:
                    code |= set(check)        

        # 技術者IDを取得し、重複しない場合はengineer_lits_idsに格納する。
            if code :
                for code in code:
                    check = Engineerinfo.objects.filter(name__icontains=code.code).values_list('id',flat=True).distinct()
                    if engineer_list_ids is None:
                        engineer_list_ids = set(check)
                    else:
                        engineer_list_ids |= set(check)   
        
        #　案件情報の必須スキルを持っている技術者の情報の取得
        if engineer_list_ids != None:
            for engineer_list_ids in engineer_list_ids:
                engineerlist = Engineerinfo.objects.filter(id__icontains=engineer_list_ids)
                engineer_list.append(engineerlist)   
        
        # contextに文字を入れることで送る
        context = {'project' : project,
                'engineer' : engineer_list
                }
        return context

#######################################
#技術者登録処理
#######################################
def engineer_upload(request):
    if request.method == 'POST':
        #横浜開発センター用
        if 'testfile1' in request.FILES:
            # スキルシート下部の技術経験　項番MAX30行まで指定
            # count = 30 * 4行
            count = 120


            # 画面上で受け取ったファイルを受け取る
            if request.method == 'POST' and request.FILES['testfile1']:
                myfile = request.FILES['testfile1']
                
                #　Excel以外のファイルが渡されたときにエラー文を出力するための処理
                if not myfile.name.endswith(".xlsx"):
                    error = "Excelファイル以外のファイルが選択されました。"
                    context = {"error" : error
                    }
                    return render(request, 'search/engineer_file_upload.html', context)

                fs = FileSystemStorage()
                fs.save(myfile.name, myfile)

                try:
                    # ワークブックを開く
                    book = excel.load_workbook(myfile, data_only=True)
                    # 技術者情報一覧.xlsxの最初(0番目)のシートを取り出す
                    sheet = book.worksheets[0]
                    # 名前・年齢・性別・最寄駅・最終学歴のデータを取り出す
                    name = sheet.cell(8,3).value
                    age = sheet.cell(7,33).value
                    gender = sheet.cell(9, 24).value
                    if gender == "女":
                        gender = 0
                    elif gender == "男":
                        gender = 1    
                    neareststation = sheet.cell(11, 33).value
                    finaleducation = sheet.cell(11, 3).value
                    # 各変数をDBに格納
                    insert = Engineerinfo(name=name, age=age, gender=gender, neareststation=neareststation, finaleducation=finaleducation)
                    insert.save()
                except:
                    return render(request, 'search/error.html')

                try:
                    year = None
                    month = None

                    os = []
                    oslevel = []
                    osperiod = []
                    for i in range(9):
                        # osの値をosのリストに格納
                        if sheet.cell(18+i,3).value != None:
                            os.append(sheet.cell(18+i,3).value)
                            # osのスキルレベルをoslevelのリストに格納
                            for j in range(5):
                                if sheet.cell(18+i, 12+j).value != None:
                                    oslevel.append(5-j)
                                    break
                                elif j == 4:
                                    oslevel.append("")
                            # osの経験年数をyearとmonthに格納        
                            if sheet.cell(18+i, 17).value != None:
                                year = sheet.cell(18+i, 17).value
                            if sheet.cell(18+i, 19).value != None:
                                month = sheet.cell(18+i, 19).value
                            # yearとmonthの値を結合してosperiodのリストに格納
                            if year != None and month != None:
                                osperiod.append(str(year) + "年" + str(month) + "月")
                            elif year == None and month != None: 
                                osperiod.append(str(month) + "月")
                            elif year != None and month == None: 
                                osperiod.append(str(year) + "年")
                            else:
                                osperiod.append("")    
                            year = None
                            month = None

                    server = []
                    serverlevel = []
                    serverperiod = []
                    for i in range(9):
                        # serverの値をserverのリストに格納
                        if sheet.cell(28+i,3).value != None:
                            server.append(sheet.cell(28+i, 3).value)
                            # serverのスキルレベルをserverlevelのリストに格納
                            for j in range(5):
                                if sheet.cell(28+i, 12+j).value != None:
                                    serverlevel.append(5-j)
                                    break
                                elif j == 4:
                                    serverlevel.append("")    
                            # serverの経験年数をyearとmonthに格納         
                            if sheet.cell(28+i, 17).value != None:
                                year = sheet.cell(28+i, 17).value
                            if sheet.cell(28+i, 19).value != None:
                                month = sheet.cell(28+i, 19).value
                            # yearとmonthの値を結合してserverperiodのリストに格納    
                            if year != None and month != None:
                                serverperiod.append(str(year) + "年" + str(month) + "月")
                            elif year == None and month != None: 
                                serverperiod.append(str(month) + "月")
                            elif year != None and month == None: 
                                serverperiod.append(str(year) + "年")
                            else:
                                serverperiod.append("")        
                            year = None
                            month = None        

                    # DBのみレベルが4段階(A～D)
                    DB = []
                    DBlevel = []
                    DBperiod = []
                    for i in range(9):
                        # DBの値をDBのリストに格納
                        if sheet.cell(47+i,3).value != None:
                            DB.append(sheet.cell(47+i, 3).value)
                            # DBのスキルレベルをDBlevelのリストに格納
                            for j in range(5):
                                if sheet.cell(47+i, 13+j).value is not None:
                                    if j == 0:
                                        DBlevel.append("A")
                                        break
                                    elif j == 1:
                                        DBlevel.append("B")
                                        break
                                    elif j == 2:
                                        DBlevel.append("C")
                                        break
                                    elif j == 3:
                                        DBlevel.append("D")
                                        break
                                    elif j == 4:
                                        DBlevel.append("")
                                        break
                                elif j == 4:
                                    DBlevel.append("")     
                                      
                            # DBの経験年数をyearとmonthに格納             
                            if sheet.cell(47+i, 17).value != None:
                                year = sheet.cell(47+i, 17).value
                            if sheet.cell(47+i, 19).value != None:
                                month = sheet.cell(47+i, 19).value
                            # yearとmonthの値を結合してDBperiodのリストに格納
                            if year != None and month != None:
                                DBperiod.append(str(year) + "年" + str(month) + "月")
                            elif year == None and month != None: 
                                DBperiod.append(str(month) + "月")
                            elif year != None and month == None: 
                                DBperiod.append(str(year) + "年")
                            else:
                                DBperiod.append("")     
                            year = None
                            month = None   

                    language = []
                    langagelevel = []
                    languageperiod = []
                    for i in range(9):
                        # languageの値をlanguageのリストに格納
                        if sheet.cell(18+i,21).value != None:
                            language.append(sheet.cell(18+i, 21).value)
                            # languageのスキルレベルをlanguagelevelのリストに格納
                            for j in range(5):
                                if sheet.cell(18+i, 32+j).value != None:
                                    langagelevel.append(5-j)
                                    break
                                elif j == 4:
                                    langagelevel.append("")
                            # languageの経験年数をyearとmonthに格納        
                            if sheet.cell(18+i, 37).value != None:
                                year = sheet.cell(18+i, 37).value
                            if sheet.cell(18+i, 39).value != None:
                                month = sheet.cell(18+i, 39).value
                            # yearとmonthの値を結合してlanguageperiodのリストに格納
                            if year != None and month != None:
                                languageperiod.append(str(year) + "年" + str(month) + "月")
                            elif year == None and month != None: 
                                languageperiod.append(str(month) + "月")
                            elif year != None and month == None: 
                                languageperiod.append(str(year) + "年")
                            else:
                                languageperiod.append("")    
                            year = None
                            month = None
                            
                    tool = []
                    toollevel = []
                    toolperiod = []
                    for i in range(7):
                        # toolの値をtoolのリストに格納
                        if sheet.cell(28+i,21).value != None:
                            tool.append(sheet.cell(28+i, 21).value)
                            # toolのスキルレベルをtoollevelのリストに格納
                            for j in range(5):
                                if sheet.cell(28+i, 32+j).value != None:
                                    toollevel.append(5-j)
                                    break
                                elif j == 4:
                                    toollevel.append("")    
                            # toolの経験年数をyearとmonthに格納         
                            if sheet.cell(28+i, 37).value != None:
                                year = sheet.cell(28+i, 37).value
                            if sheet.cell(28+i, 39).value != None:
                                month = sheet.cell(28+i, 39).value
                            # yearとmonthの値を結合してtoolperiodのリストに格納
                            if year != None and month != None:
                                toolperiod.append(str(year) + "年" + str(month) + "月")
                            elif year == None and month != None: 
                                toolperiod.append(str(month) + "月")
                            elif year != None and month == None: 
                                toolperiod.append(str(year) + "年")
                            else:
                                toolperiod.append("")
                            year = None
                            month = None        

                    qualification = []
                    qualificationperiod = [] 
                    for i in range(4):
                        # qualificationの値をqualificationのリストに格納
                        if sheet.cell(52+i,21).value != None:
                            qualification.append(sheet.cell(52+i, 21).value)
                            # toolの取得年月をyearとmonthに格納
                            if sheet.cell(52+i, 36).value != None:
                                year = sheet.cell(52+i, 36).value
                            if sheet.cell(52+i, 39).value != None:
                                month = sheet.cell(52+i, 39).value   
                            # yearとmonthの値を結合してqualificationperiodのリストに格納
                            if year != None and month != None:
                                qualificationperiod.append(str(year) + "年" + str(month) + "月")
                            elif year == None and month != None: 
                                qualificationperiod.append(str(month) + "月")
                            elif year != None and month == None: 
                                qualificationperiod.append(str(year) + "年")
                            else:
                                qualificationperiod.append("")   
                            year = None
                            month = None 

                    experience = []
                    # experienceの中にcount(最上部で定義した)回数分for文で回す
                    # 要件定義などの経験の欄に黒い丸がある時experienceの変数に格納
                    for i in range(count):
                        for j in range(8):
                            if sheet.cell(61+i, 20+j).value != None and j == 0:
                                    experience.append("要件定義")
                            elif sheet.cell(61+i, 20+j).value != None and j == 1:
                                    experience.append("基本設計")
                            elif sheet.cell(61+i, 20+j).value != None and j == 2:
                                    experience.append("詳細設計")
                            elif sheet.cell(61+i, 20+j).value != None and j == 3:
                                    experience.append("製造")
                            elif sheet.cell(61+i, 20+j).value != None and j == 4:
                                    experience.append("単体試験")
                            elif sheet.cell(61+i, 20+j).value != None and j == 5:
                                    experience.append("結合試験")
                            elif sheet.cell(61+i, 20+j).value != None and j == 6:
                                    experience.append("総合試験")
                            elif sheet.cell(61+i, 20+j).value != None and j == 7:
                                    experience.append("運用保守")
                            experience = list(set(experience))

                    role = []
                    # roleの中にcount(最上部で定義した)回数分for文で回す
                    # PMなどの役割をroleの変数に格納
                    for i in range (count):
                        if sheet.cell(61+i, 18).value != None:
                            check = sheet.cell(61+i, 18).value
                            if "PM" in check:
                                role.append("PM")
                            if "PL" in check:
                                role.append("PL")
                            if "GL" in check:
                                role.append("GL")
                            if "SE" in check:
                                role.append("SE")
                            if "PG" in check:
                                role.append("PG")
                            if "T" in check:
                                role.append("T")
                            if "OP" in check:
                                role.append("OP")
                            role = list(set(role))

                    # 名前と技術者情報テーブルで登録した名前が一致するidを取得
                    id = Engineerinfo.objects.filter(name=name, age=age, gender=gender, neareststation=neareststation, finaleducation=finaleducation).values_list('id', flat=True)
                    id = list(set(id))

                    # 資格の文字列の中の空白を削除
                    # コードマスタの中から一致する資格のskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(qualification)):
                        qualification[i] = qualification[i].replace(' ', '')
                        qualification[i] = qualification[i].replace('　', '')
                        skillid =  Codemaster.objects.filter(name__iexact=qualification[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])
                            
                            # 資格に関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=qualification[i], years=qualificationperiod[i])
                            insert.save()

                    # 言語の文字列の中の空白を削除
                    # コードマスタの中から一致する言語のskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(language)):
                        language[i] = language[i].replace(' ', '')
                        language[i] = language[i].replace('　', '')
                        skillid =  Codemaster.objects.filter(name__iexact=language[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])
                            # 言語に関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=language[i], years=languageperiod[i], skilllevel=langagelevel[i])
                            insert.save()

                    # ツールの文字列の中の空白を削除
                    # コードマスタの中から一致するツールのskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(tool)):
                        tool[i] = tool[i].replace(' ', '')
                        tool[i] = tool[i].replace('　', '')
                        skillid =  Codemaster.objects.filter(name__iexact=tool[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])
                            
                            # ツールに関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=tool[i], years=toolperiod[i], skilllevel=toollevel[i])
                            insert.save()
                    
                    # osの文字列の中の空白を削除
                    # コードマスタの中から一致するosのskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(os)):
                        os[i] = os[i].replace(' ', '')
                        os[i] = os[i].replace('　', '')
                        skillid =  Codemaster.objects.filter(name__iexact=os[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])
                            
                            # osに関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=os[i], years=osperiod[i], skilllevel=oslevel[i])
                            insert.save()

                    # サーバーの文字列の中の空白を削除
                    # コードマスタの中から一致するサーバーのskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(server)):
                        server[i] = server[i].replace(' ', '')
                        server[i] = server[i].replace('　', '')
                        skillid =  Codemaster.objects.filter(name__iexact=server[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])
                            
                            # サーバーに関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=server[i], years=serverperiod[i], skilllevel=serverlevel[i])
                            insert.save()

                    # DBの文字列の中の空白を削除
                    # コードマスタの中から一致するDBのskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(DB)):
                        DB[i] = DB[i].replace(' ', '')
                        DB[i] = DB[i].replace('　', '')
                        skillid =  Codemaster.objects.filter(name__iexact=DB[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])
                            
                            # DBに関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=DB[i], years=DBperiod[i], skilllevel=DBlevel[i])
                            insert.save()

                    # 技術経験(PMなど)の文字列の中の空白を削除
                    # コードマスタの中から一致する技術経験のskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(experience)):
                        experience[i] = experience[i].replace(' ', '')
                        experience[i] = experience[i].replace('　', '')
                        skillid =  Codemaster.objects.filter(name__iexact=experience[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])

                            # 技術経験(PMなど)に関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=experience[i])
                            insert.save()

                    # 役割の文字列の中の空白を削除
                    # コードマスタの中から一致する役割のskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(role)):
                        role[i] = role[i].replace(' ', '')
                        role[i] = role[i].replace('　', '')
                        skillid =  Codemaster.objects.filter(name__iexact=role[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])
                        
                            # 役割に関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=role[i])
                            insert.save()

                    # ロードしたExcelファイルを閉じる
                    book.close()
                    complete = {"complete" : True
                    }
                    return render(request, 'search/engineer_file_upload.html',complete)                

                # エラー時に技術者情報テーブルに登録している名前のデータと登録処理中のスキル情報テーブルのデータを削除
                except:
                    Engineerinfo.objects.filter(name=name).delete()
                    Skillinfo.objects.filter(id=id[0]).delete()
                    return render(request) 
        
            return render(request, 'search/engineer_file_upload.html')
        
        elif 'testfile2' in request.FILES:

            # スキルシート下部の技術経験　項番MAX30行まで指定
            # count = 30 * 9行
            count = 270

        
            # 画面上で受け取ったファイルを受け取る
            if request.method == 'POST' and request.FILES['testfile2']:
                myfile = request.FILES['testfile2']
                
                #　Excel以外のファイルが渡されたときにエラー文を出力するための処理
                if not myfile.name.endswith(".xlsx"):
                    error = "Excelファイル以外のファイルが選択されました。"
                    context = {"error" : error
                    }
                    return render(request, 'search/engineer_file_upload.html', context)

                fs = FileSystemStorage()
                fs.save(myfile.name, myfile)

                try:
                    #ワークブックを開く
                    book = excel.load_workbook(myfile, data_only=True)
                    #技術者情報一覧.xlsxの2つ目(1番目)のシートを取り出す
                    sheet_1 = book.worksheets[1]

                    # 名前・年齢・最寄駅・最終学歴のデータを取り出す
                    name1 = sheet_1.cell(6,9).value
                    name2 = sheet_1.cell(6,16).value
                    # 名前は苗字と名前を結合
                    if name1 and name2:
                        name = str(name1) + str(name2)
                    elif name1:
                        name = str(name1)
                    elif name2:
                        name = str(name2)
                    age = sheet_1.cell(7,25).value
                    neareststation = sheet_1.cell(6, 40).value
                    finaleducation = sheet_1.cell(8, 9).value

                    # 技術者情報をDBに登録
                    insert = Engineerinfo(name=name, age=age, neareststation=neareststation, finaleducation=finaleducation)
                    insert.save()

                except:
                    return render(request, 'search/error.html')

                try:
                    #ワークブックを開く
                    book = excel.load_workbook(myfile, data_only=True)
                    #技術者情報一覧.xlsxの3つ目(2番目)のシートを取り出す
                    sheet_2 = book.worksheets[2]

                    year = None
                    month = None

                    os = []
                    oslevel = []
                    osperiod = []
                    for i in range(9):
                        # osの値をosのリストに格納
                        if sheet_2.cell(7+i,2).value != None:
                            os.append(sheet_2.cell(7+i,2).value)
                             # osのスキルレベルをoslevelのリストに格納
                            for j in range(5):
                                if sheet_2.cell(7+i, 16+j).value != None:
                                    oslevel.append(5-j)
                                    break
                                elif j == 4:
                                    oslevel.append("")
                            # osの経験年数をyearとmonthに格納        
                            if sheet_2.cell(7+i, 21).value != None:
                                year = sheet_2.cell(7+i, 21).value
                            if sheet_2.cell(7+i, 23).value != None:
                                month = sheet_2.cell(7+i, 23).value
                            # yearとmonthの値を結合してosperiodのリストに格納
                            if year != None and month != None:
                                osperiod.append(str(year) + "年" + str(month) + "月")
                            elif year == None and month != None: 
                                osperiod.append(str(month) + "月")
                            elif year != None and month == None: 
                                osperiod.append(str(year) + "年")
                            else:
                                osperiod.append("")    
                            year = None
                            month = None

                    server = []
                    serverlevel = []
                    serverperiod = []
                    for i in range(9):
                        # サーバーの値をserverのリストに格納
                        if sheet_2.cell(17+i,2).value != None:
                            server.append(sheet_2.cell(17+i, 2).value)
                            # サーバーのスキルレベルをserverlevelのリストに格納
                            for j in range(5):
                                if sheet_2.cell(17+i, 16+j).value != None:
                                    serverlevel.append(5-j)
                                    break
                                elif j == 4:
                                    serverlevel.append("")
                            # サーバーの経験年数をyearとmonthに格納        
                            if sheet_2.cell(17+i, 21).value != None:
                                year = sheet_2.cell(17+i, 21).value
                            if sheet_2.cell(17+i, 23).value != None:
                                month = sheet_2.cell(17+i, 23).value
                            # yearとmonthの値を結合してserverperiodのリストに格納
                            if year != None and month != None:
                                serverperiod.append(str(year) + "年" + str(month) + "月")
                            elif year == None and month != None: 
                                serverperiod.append(str(month) + "月")
                            elif year != None and month == None: 
                                serverperiod.append(str(year) + "年")
                            else:
                                serverperiod.append("")      
                            year = None
                            month = None
                      
                    # DBのみレベルが4段階(A～D)
                    DB = []
                    DBlevel = []
                    DBperiod = []
                    for i in range(9):
                        # DBの値をDBのリストに格納
                        if sheet_2.cell(36+i,2).value != None:
                            DB.append(sheet_2.cell(36+i, 2).value)
                            # DBのスキルレベルをDBlevelのリストに格納
                            for j in range(5):
                                if sheet_2.cell(36+i, 17+j).value is not None:
                                    if j == 0:
                                        DBlevel.append("A")
                                        break
                                    elif j == 1:
                                        DBlevel.append("B")
                                        break
                                    elif j == 2:
                                        DBlevel.append("C")
                                        break
                                    elif j == 3:
                                        DBlevel.append("D")
                                        break
                                    elif j == 4:
                                        DBlevel.append("")
                                        break
                                elif j == 4:
                                    DBlevel.append("")        
                            # DBの経験年数をyearとmonthに格納  
                            if sheet_2.cell(36+i, 21).value != None:
                                year = sheet_2.cell(36+i, 21).value
                            if sheet_2.cell(36+i, 23).value != None:
                                month = sheet_2.cell(36+i, 23).value
                            # yearとmonthの値を結合してDBperiodのリストに格納
                            if year != None and month != None:
                                DBperiod.append(str(year) + "年" + str(month) + "月")
                            elif year == None and month != None: 
                                DBperiod.append(str(month) + "月")
                            elif year != None and month == None: 
                                DBperiod.append(str(year) + "年")
                            else:
                                DBperiod.append("")      
                            year = None
                            month = None

                    language = []
                    languagelevel = []
                    languageperiod = []
                    for i in range(10):
                        # 言語の値をlanguageのリストに格納
                        if sheet_2.cell(7+i,25).value != None:
                            language.append(sheet_2.cell(7+i, 25).value)
                            # 言語のスキルレベルをlanguagelevelのリストに格納
                            for j in range(5):
                                if sheet_2.cell(7+i, 40+j).value != None:
                                    languagelevel.append(5-j)
                                    break
                                elif j == 4:
                                    languagelevel.append("")     
                            # 言語の経験年数をyearとmonthに格納
                            if sheet_2.cell(7+i, 45).value != None:
                                year = sheet_2.cell(7+i, 45).value
                            if sheet_2.cell(7+i, 47).value != None:
                                month = sheet_2.cell(7+i, 47).value
                            # yearとmonthの値を結合してDBperiodのリストに格納
                            if year != None and month != None:
                                languageperiod.append(str(year) + "年" + str(month) + "月")
                            elif year == None and month != None: 
                                languageperiod.append(str(month) + "月")
                            elif year != None and month == None: 
                                languageperiod.append(str(year) + "年")
                            else:
                                languageperiod.append("")     
                            year = None
                            month = None
                            
                    tool = []
                    toollevel = []
                    toolperiod = []
                    for i in range(3):
                        # ツールの値をlanguageのリストに格納
                        if sheet_2.cell(18+i,25).value != None:
                            tool.append(sheet_2.cell(18+i, 25).value)
                            # ツールのスキルレベルをlanguagelevelのリストに格納
                            for j in range(5):
                                if sheet_2.cell(18+i, 40+j).value != None:
                                    toollevel.append(5-j)
                                    break
                                elif j == 4:
                                    toollevel.append("")    
                            # ツールの経験年数をyearとmonthに格納
                            if sheet_2.cell(18+i, 45).value != None:
                                year = sheet_2.cell(18+i, 45).value
                            if sheet_2.cell(18+i, 47).value != None:
                                month = sheet_2.cell(18+i, 47).value
                            # yearとmonthの値を結合してDBperiodのリストに格納
                            if year != None and month != None:
                                toolperiod.append(str(year) + "年" + str(month) + "月")
                            elif year == None and month != None: 
                                toolperiod.append(str(month) + "月")
                            elif year != None and month == None: 
                                toolperiod.append(str(year) + "年")
                            else:
                                toolperiod.append("")    
                            year = None
                            month = None

                    qualification = []
                    for i in range(6):
                        # 資格の値をlanguageのリストに格納
                        if sheet_1.cell(8+i,30).value != None:
                            qualification.append(sheet_1.cell(8+i, 30).value)

                    experience = []
                    # experienceの中にcount(最上部で定義した)回数分for文で回す
                    # 要件定義などの担当経験の欄に黒い丸がある時experienceの変数に格納
                    for i in range(count):
                            if sheet_1.cell(18+i, 37).value == "●":
                                if sheet_1.cell(16+i, 37).value == "要件定義":
                                    experience.append("要件定義")
                                elif sheet_1.cell(16+i, 37).value == "製造":
                                    experience.append("製造")
                                elif sheet_1.cell(16+i, 37).value == "結合試験":    
                                    experience.append("結合試験")
                            if sheet_1.cell(18+i, 42).value == "●":
                                if sheet_1.cell(16+i, 42).value == "基本設計":
                                    experience.append("基本設計")
                                elif sheet_1.cell(16+i, 42).value == "テスト設計":    
                                    experience.append("テスト設計")
                                elif sheet_1.cell(16+i, 42).value == "総合試験":
                                    experience.append("総合試験")
                            if sheet_1.cell(18+i, 47).value == "●" :
                                if sheet_1.cell(16+i, 47).value == "詳細設計":
                                    experience.append("詳細設計")
                                elif sheet_1.cell(16+i, 47).value == "単体試験":    
                                    experience.append("単体試験")
                                elif sheet_1.cell(16+i, 47).value == "運用保守":     
                                    experience.append("運用保守")
                            experience = list(set(experience))

                    role = []
                    # roleの中にcount(最上部で定義した)回数分for文で回す
                    # PMなどの役割をroleの変数に格納
                    for i in range (count):
                        num = 9*i
                        if sheet_1.cell(16+num, 24).value != None:
                            check = sheet_1.cell(16+num, 24).value
                            if "PM" in check:
                                role.append("PM")
                            if "PL" in check:
                                role.append("PL")
                            if "GL" in check:
                                role.append("GL")
                            if "SE" in check:
                                role.append("SE")
                            if "PG" in check:
                                role.append("PG")
                            if "ﾃｽﾀｰ" in check:
                                role.append("T")
                            if "OP" in check:
                                role.append("OP")
                            role = list(set(role))

                    # 名前と技術者情報テーブルで登録した名前が一致するidを取得
                    id = Engineerinfo.objects.filter(name=name).values_list('id', flat=True)
                    id = list(set(id))

                    # osの文字列の中の空白を削除
                    # コードマスタの中から一致するosのskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(os)):
                        os[i] = os[i].replace(' ', '')
                        os[i] = os[i].replace('　', '')
                        skillid =  Codemaster.objects.filter(name__iexact=os[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])
                            
                            # osに関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=os[i], years=osperiod[i], skilllevel=oslevel[i])
                            insert.save()

                    # サーバーの文字列の中の空白を削除
                    # コードマスタの中から一致するサーバーのskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(server)):
                        server[i] = server[i].replace(' ', '')
                        server[i] = server[i].replace('　', '')
                        skillid =  Codemaster.objects.filter(name__iexact=server[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])
                           
                            # サーバーに関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=server[i], years=serverperiod[i], skilllevel=serverlevel[i])
                            insert.save()

                    # DBの文字列の中の空白を削除
                    # コードマスタの中から一致するDBのskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(DB)):
                        DB[i] = DB[i].replace(' ', '')
                        DB[i] = DB[i].replace('　', '')
                        skillid =  Codemaster.objects.filter(name__iexact=DB[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])
                            
                            # DBに関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=DB[i], years=DBperiod[i], skilllevel=DBlevel[i])
                            insert.save()

                    # 言語の文字列の中の空白を削除
                    # コードマスタの中から一致する言語のskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(language)):
                        language[i] = language[i].replace(' ', '')
                        language[i] = language[i].replace('　', '')
                        skillid =  Codemaster.objects.filter(name__iexact=language[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])
                            
                            # 言語に関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=language[i], years=languageperiod[i], skilllevel=languagelevel[i])
                            insert.save()

                    # ツールの文字列の中の空白を削除
                    # コードマスタの中から一致するツールのskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(tool)):
                        tool[i] = tool[i].replace(' ', '')
                        tool[i] = tool[i].replace('　', '')
                        skillid =  Codemaster.objects.filter(name__iexact=tool[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])
                            
                            # ツールに関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=tool[i], years=toolperiod[i], skilllevel=toollevel[i])
                            insert.save()                

                    # 資格の文字列の中の空白を削除
                    # コードマスタの中から一致する資格のskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(qualification)):
                        qualification[i] = qualification[i].replace(' ', '')
                        qualification[i] = qualification[i].replace('　', '')
                        skillid =  Codemaster.objects.filter(name__iexact=qualification[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])
                            
                            # 資格に関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=qualification[i])
                            insert.save()

                    # 経験の文字列の中の空白を削除
                    # コードマスタの中から一致する役割のskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(experience)):
                        skillid =  Codemaster.objects.filter(name=experience[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])
                            
                            # 役割に関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=experience[i])
                            insert.save()  

                    # 役割の文字列の中の空白を削除
                    # コードマスタの中から一致する役割のskillidを取得し、idとskillidでskillinfoidを作成
                    for i in range(len(role)):
                        skillid =  Codemaster.objects.filter(name=role[i]).values_list('skillid', flat=True)
                        skillid = list(set(skillid))
                        if skillid:
                            skillinfoid = str(id[0]) + str(skillid[0])
                            
                            # 役割に関する情報をスキル情報テーブルに登録
                            insert = Skillinfo(id=id[0], code=name, skillid=skillid[0], skillinfoid=skillinfoid, classification=role[i])
                            insert.save()    

                    # ロードしたExcelファイルを閉じる
                    book.close()     
                    complete = {"complete" : True
                    }     
                    return render(request, 'search/engineer_file_upload.html',complete)

                # エラー時に技術者情報テーブルに登録している名前のデータと登録処理中のスキル情報テーブルのデータを削除         
                except:
                    Engineerinfo.objects.filter(name=name).delete()
                    Skillinfo.objects.filter(id=id[0]).delete()
                    return render(request, 'search/error.html')
                
    return render(request, 'search/engineer_file_upload.html')

#######################################
#エラー処理
#######################################
def error(request):
     return render(request, 'search/error.html')